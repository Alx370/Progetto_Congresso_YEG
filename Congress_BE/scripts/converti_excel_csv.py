from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
import argparse
import csv

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR.parent.parent / "Dataset Evento Congresso 2025.xlsx"
DEFAULT_PARTECIPANTI = SCRIPT_DIR.parent / "dati" / "partecipanti.csv"
DEFAULT_INTERAZIONI = SCRIPT_DIR.parent / "dati" / "interazioni.csv"


def leggi_dizionario(workbook):
    dizionario = []
    foglio = workbook["01_Interazioni"]

    for riga in foglio.iter_rows(min_row=2, values_only=True):
        colonna, intestazione, nome_tecnico, tipo_dato, fase, descrizione = riga
        indice = column_index_from_string(colonna) - 1

        if indice >= 7:
            dizionario.append({
                "indice": indice,
                "intestazione": intestazione,
                "nome_tecnico": nome_tecnico,
                "tipo_dato": tipo_dato,
                "fase": fase,
                "descrizione": descrizione,
            })

    return dizionario


def normalizza_booleano(valore):
    if valore in (1, True, "1", "true", "vero", "si", "sì"):
        return "true"
    if valore in (0, False, "0", "false", "falso", "no"):
        return "false"
    raise ValueError(f"Valore booleano non valido: {valore}")


def normalizza_valore(valore, tipo_dato):
    if valore is None or str(valore).strip() == "":
        return ""

    if tipo_dato == "booleano":
        return normalizza_booleano(valore)

    if tipo_dato == "data":
        if isinstance(valore, (date, datetime)):
            return valore.strftime("%Y-%m-%d")
        return datetime.strptime(str(valore), "%d/%m/%Y").strftime("%Y-%m-%d")

    if tipo_dato in ("conteggio", "minuti"):
        return str(int(valore))

    if tipo_dato == "tasso da 0 a 1":
        return str(Decimal(str(valore)).quantize(Decimal("0.0001")))

    return str(valore).strip()


def converti(percorso_excel: Path, partecipanti_csv: Path, interazioni_csv: Path):
    workbook = load_workbook(percorso_excel, read_only=True, data_only=True)
    dizionario = leggi_dizionario(workbook)
    foglio = workbook["02_Partecipanti"]

    intestazioni = [cella.value for cella in next(foglio.iter_rows(min_row=1, max_row=1))]
    for definizione in dizionario:
        trovata = intestazioni[definizione["indice"]]
        if trovata != definizione["intestazione"]:
            raise ValueError(
                f"Intestazione non valida: attesa '{definizione['intestazione']}', trovata '{trovata}'"
            )

    partecipanti_csv.parent.mkdir(parents=True, exist_ok=True)
    interazioni_csv.parent.mkdir(parents=True, exist_ok=True)

    totale = 0
    with partecipanti_csv.open("w", encoding="utf-8", newline="") as file_partecipanti, \
            interazioni_csv.open("w", encoding="utf-8", newline="") as file_interazioni:

        writer_partecipanti = csv.writer(file_partecipanti, delimiter=";")
        writer_interazioni = csv.writer(file_interazioni, delimiter=";")

        writer_partecipanti.writerow(intestazioni[:7])
        writer_interazioni.writerow(
            ["email"] + [definizione["nome_tecnico"] for definizione in dizionario]
        )

        for riga in foglio.iter_rows(min_row=2, values_only=True):
            partecipante = list(riga[:7])
            partecipante[6] = normalizza_booleano(partecipante[6])

            interazioni = [str(riga[2]).strip().lower()]
            for definizione in dizionario:
                interazioni.append(normalizza_valore(
                    riga[definizione["indice"]],
                    definizione["tipo_dato"]
                ))

            writer_partecipanti.writerow(partecipante)
            writer_interazioni.writerow(interazioni)
            totale += 1

    workbook.close()
    return totale


def main():
    parser = argparse.ArgumentParser(description="Converte il dataset Excel in due CSV")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--partecipanti", type=Path, default=DEFAULT_PARTECIPANTI)
    parser.add_argument("--interazioni", type=Path, default=DEFAULT_INTERAZIONI)
    argomenti = parser.parse_args()

    totale = converti(argomenti.input, argomenti.partecipanti, argomenti.interazioni)
    print(f"Creati partecipanti.csv e interazioni.csv con {totale} partecipanti")


if __name__ == "__main__":
    main()
